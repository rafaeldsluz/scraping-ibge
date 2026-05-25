"""
exporter.py — Exportação dos resultados para planilha Excel formatada.

Gera duas abas:
  "Resultado" — todos os registros com MAPA_URL e STATUS
  "Sumário"   — estatísticas consolidadas por status e por município
"""

import logging
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

# Ordem das colunas na planilha de saída
OUTPUT_COLUMNS = [
    "CD_SETOR",
    "CD_MUN",
    "NM_MUN",
    "CD_DIST",
    "NM_DIST",
    "CD_SUBDIST",
    "NM_SUBDIST",
    "TIPO_AREA",
    "MAPA_URL",
    "STATUS",
]

# Paleta de cores para formatação do cabeçalho
_HEADER_BG = "1F4E79"   # azul escuro IBGE
_HEADER_FG = "FFFFFF"   # branco


def save_results(
    records:     list[dict],
    output_path: Path,
    logger:      logging.Logger,
) -> None:
    """
    Salva os resultados em planilha Excel com formatação profissional.

    Args:
        records:     lista de dicts com chaves de OUTPUT_COLUMNS
        output_path: caminho do arquivo de saída (.xlsx)
        logger:      logger do projeto
    """
    if not records:
        logger.warning("  Nenhum registro para exportar.")
        return

    df = pd.DataFrame(records, columns=OUTPUT_COLUMNS)

    total      = len(df)
    found      = (df["STATUS"] == "ENCONTRADO").sum()
    not_found  = total - found
    logger.info(
        f"  {total:,} registros  |  "
        f"{found:,} encontrados  |  "
        f"{not_found:,} não encontrados"
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # ── Aba principal ─────────────────────────────────────────────
        df.to_excel(writer, index=False, sheet_name="Resultado")
        _format_sheet(writer.sheets["Resultado"], freeze="A2")

        # ── Aba de sumário ────────────────────────────────────────────
        _write_summary(writer, df)

    logger.info(f"  Arquivo salvo: {output_path}")


# ── Formatação ────────────────────────────────────────────────────────────────

def _format_sheet(ws, freeze: str = "A2") -> None:
    """Aplica estilo corporativo à aba: cabeçalho colorido, colunas ajustadas."""
    header_fill = PatternFill(
        start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid"
    )
    header_font = Font(name="Calibri", bold=True, color=_HEADER_FG, size=11)
    data_font   = Font(name="Calibri", size=10)

    # Formata cabeçalho (linha 1)
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    ws.row_dimensions[1].height = 30

    # Fonte das linhas de dados
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font      = data_font
            cell.alignment = Alignment(vertical="center")

    # Ajuste automático de largura das colunas
    for col_cells in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col_cells)
        col_letter = col_cells[0].column_letter
        ws.column_dimensions[col_letter].width = min(max_len + 4, 70)

    ws.freeze_panes  = freeze
    ws.auto_filter.ref = ws.dimensions


def _write_summary(writer: pd.ExcelWriter, df: pd.DataFrame) -> None:
    """Cria aba 'Sumário' com totais por status e por município."""
    rows_status = [
        ("Total de setores",            len(df)),
        ("Mapas encontrados",           (df["STATUS"] == "ENCONTRADO").sum()),
        ("Município não encontrado",    (df["STATUS"] == "MUNICÍPIO NÃO ENCONTRADO").sum()),
        ("Distrito não encontrado",     (df["STATUS"] == "DISTRITO NÃO ENCONTRADO").sum()),
        ("Setor não encontrado",        (df["STATUS"] == "SETOR NÃO ENCONTRADO").sum()),
        ("Erro de requisição",          (df["STATUS"] == "ERRO DE REQUISIÇÃO").sum()),
        ("─" * 30,                      ""),
        ("Municípios únicos",           df["CD_MUN"].nunique()),
        ("Distritos únicos",            df["CD_DIST"].nunique()),
        ("Subdistritos únicos",         df["CD_SUBDIST"].nunique()),
        ("Setores em área urbana (MSU)", (df["TIPO_AREA"] == "MSU").sum()),
        ("Setores em área rural (MSR)",  (df["TIPO_AREA"] == "MSR").sum()),
    ]

    summary_df = pd.DataFrame(rows_status, columns=["Métrica", "Valor"])
    summary_df.to_excel(writer, index=False, sheet_name="Sumário")
    _format_sheet(writer.sheets["Sumário"], freeze="A2")

    # ── Totais por município ──────────────────────────────────────────
    mun_stats = (
        df.groupby(["CD_MUN", "NM_MUN"])
        .agg(
            Total=("CD_SETOR", "count"),
            Encontrados=("STATUS", lambda s: (s == "ENCONTRADO").sum()),
        )
        .reset_index()
    )
    mun_stats["Taxa (%)"] = (
        mun_stats["Encontrados"] / mun_stats["Total"] * 100
    ).round(1)

    mun_stats.to_excel(
        writer, index=False, sheet_name="Por Município", startrow=0
    )
    _format_sheet(writer.sheets["Por Município"], freeze="A2")
